"""Historical odds backfill from the aussportsbetting.com NRL workbook.

Free for personal use; updated within ~a day of each round. Provides
open/min/max/close for head-to-head, line, and totals markets from 2009
(2013+ for line/totals). Games are matched to feed_cache_fixtures on
(home team, away team, local date +-1 day).

Write policy:
- odds_history gets an 'open' and a 'close' snapshot per game (full detail);
- feed_cache_fixtures H2H/line columns are filled ONLY where NULL, using
  CLOSING odds (the feed's own odds were last-refresh pre-kickoff values and
  the live provider snapshot is taken inside the 6 h send gate, so closing is
  the consistent training analogue);
- the new totals columns (total_line/total_over_odds/total_under_odds) are
  NULL everywhere historically, so the COALESCE fill populates them from 2013+.
"""

from __future__ import annotations

import datetime as dt
import os
import sqlite3
import tempfile
from collections import defaultdict
from pathlib import Path

import requests

from ..nrl_data.cache_writer import update_fixture_odds
from . import store
from .team_names import canonical_team
from .validity import validated_market_values

DEFAULT_XLSX_URL = "https://www.aussportsbetting.com/historical_data/nrl.xlsx"

_HEADER_MAP = {
    "Date": "date",
    "Kick-off (local)": "kickoff_local",
    "Home Team": "home_team",
    "Away Team": "away_team",
    "Venue": "venue",
    "Home Score": "home_score",
    "Away Score": "away_score",
    "Home Odds": "h2h_home",
    "Draw Odds": "h2h_draw",
    "Away Odds": "h2h_away",
    "Home Odds Open": "h2h_home_open",
    "Home Odds Min": "h2h_home_min",
    "Home Odds Max": "h2h_home_max",
    "Home Odds Close": "h2h_home_close",
    "Away Odds Open": "h2h_away_open",
    "Away Odds Min": "h2h_away_min",
    "Away Odds Max": "h2h_away_max",
    "Away Odds Close": "h2h_away_close",
    "Home Line Open": "line_home_open",
    "Home Line Close": "line_home_close",
    "Home Line Odds Open": "line_odds_home_open",
    "Home Line Odds Close": "line_odds_home_close",
    "Away Line Odds Open": "line_odds_away_open",
    "Away Line Odds Close": "line_odds_away_close",
    "Total Score Open": "total_open",
    "Total Score Close": "total_close",
    "Total Score Over Open": "total_over_open",
    "Total Score Over Close": "total_over_close",
    "Total Score Under Open": "total_under_open",
    "Total Score Under Close": "total_under_close",
}


def download_xlsx(url: str | None = None, dest: str | Path | None = None) -> Path:
    url = url or os.environ.get("FOOTY_TIPPER_ODDS_HIST_URL") or DEFAULT_XLSX_URL
    if dest is None:
        dest = Path(tempfile.gettempdir()) / "aussportsbetting_nrl.xlsx"
    dest = Path(dest)
    response = requests.get(
        url,
        timeout=120,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 "
                "Safari/537.36"
            )
        },
    )
    response.raise_for_status()
    dest.write_bytes(response.content)
    return dest


def parse_rows(xlsx_path: str | Path) -> list[dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    header: list | None = None
    parsed: list[dict] = []
    for raw in rows_iter:
        if header is None:
            if raw and any(str(cell).strip() == "Date" for cell in raw if cell):
                header = list(raw)
            continue
        record: dict = {}
        for index, cell in enumerate(raw):
            if index >= len(header) or header[index] is None:
                continue
            key = _HEADER_MAP.get(str(header[index]).strip())
            if key:
                record[key] = cell
        if not record.get("date") or not record.get("home_team"):
            continue
        record["home_team"] = canonical_team(record.get("home_team"))
        record["away_team"] = canonical_team(record.get("away_team"))
        if not record["home_team"] or not record["away_team"]:
            continue
        parsed.append(record)
    workbook.close()
    return parsed


def _fixture_index(con: sqlite3.Connection) -> dict[tuple, list[tuple]]:
    """(home, away, local_date) -> [(game_id, year, round_id)]."""
    index: dict[tuple, list[tuple]] = defaultdict(list)
    for game_id, year, round_id, home, away, start_time in con.execute(
        "SELECT game_id, competition_year, round_id, team_home, team_away, "
        "start_time FROM feed_cache_fixtures"
    ):
        if start_time is None:
            continue
        # start_time is venue-local wall clock serialised as-if-UTC
        local_date = dt.datetime.fromtimestamp(
            float(start_time), tz=dt.timezone.utc
        ).date()
        entry = (int(float(game_id)), int(float(year)), int(float(round_id)))
        index[(home, away, local_date)].append(entry)
    return index


def _match_game(record: dict, index: dict) -> tuple | None:
    date = record["date"]
    if isinstance(date, dt.datetime):
        date = date.date()
    for offset in (0, -1, 1):
        key = (
            record["home_team"],
            record["away_team"],
            date + dt.timedelta(days=offset),
        )
        candidates = index.get(key)
        if candidates:
            return candidates[0]
    return None


def _snapshot_values(record: dict, kind: str) -> dict:
    suffix = "open" if kind == "open" else "close"
    values = {
        "h2h_odds_home": record.get(f"h2h_home_{suffix}") or record.get("h2h_home"),
        "h2h_odds_away": record.get(f"h2h_away_{suffix}") or record.get("h2h_away"),
        "h2h_odds_home_min": record.get("h2h_home_min"),
        "h2h_odds_home_max": record.get("h2h_home_max"),
        "h2h_odds_away_min": record.get("h2h_away_min"),
        "h2h_odds_away_max": record.get("h2h_away_max"),
        "line_amount_home": record.get(f"line_home_{suffix}"),
        "line_odds_home": record.get(f"line_odds_home_{suffix}"),
        "line_odds_away": record.get(f"line_odds_away_{suffix}"),
        "total_line": record.get(f"total_{suffix}"),
        "total_over_odds": record.get(f"total_over_{suffix}"),
        "total_under_odds": record.get(f"total_under_{suffix}"),
    }
    numeric_values = {
        key: float(value)
        for key, value in values.items()
        if isinstance(value, (int, float))
    }
    return validated_market_values(numeric_values)


def backfill(
    db_path: str | Path,
    url: str | None = None,
    xlsx_path: str | Path | None = None,
) -> dict:
    """Load the workbook into odds_history and fill fixture-cache gaps."""
    if xlsx_path is None:
        xlsx_path = download_xlsx(url)
    records = parse_rows(xlsx_path)

    con = sqlite3.connect(str(db_path))
    try:
        store.ensure_tables(con)
        index = _fixture_index(con)

        matched = 0
        unmatched = 0
        snapshots = 0
        fixture_updates = 0
        for record in records:
            game = _match_game(record, index)
            if game is None:
                unmatched += 1
                continue
            matched += 1
            game_id, year, round_id = game

            for kind in ("open", "close"):
                values = _snapshot_values(record, kind)
                if not values:
                    continue
                # workbook rows are per-game facts, not timestamped observations
                snapshots += store.insert_snapshot(
                    con,
                    game_id,
                    year,
                    round_id,
                    source="aussportsbetting",
                    snapshot_kind=kind,
                    snapshot_time_utc=None,
                    values=values,
                )

            close_values = _snapshot_values(record, "close")
            odds_update = {
                "team_head_to_head_odds_home": close_values.get("h2h_odds_home"),
                "team_head_to_head_odds_away": close_values.get("h2h_odds_away"),
                "team_line_amount_home": close_values.get("line_amount_home"),
                "team_line_odds_home": close_values.get("line_odds_home"),
                "team_line_odds_away": close_values.get("line_odds_away"),
                "total_line": close_values.get("total_line"),
                "total_over_odds": close_values.get("total_over_odds"),
                "total_under_odds": close_values.get("total_under_odds"),
            }
            line_home = close_values.get("line_amount_home")
            if line_home is not None:
                odds_update["team_line_amount_away"] = -line_home
            odds_update = {
                key: value for key, value in odds_update.items() if value is not None
            }
            if odds_update and update_fixture_odds(
                con, game_id, odds_update, only_when_null=True
            ):
                fixture_updates += 1

        con.commit()
        summary = {
            "records": len(records),
            "matched": matched,
            "unmatched": unmatched,
            "snapshots_inserted": snapshots,
            "fixture_rows_updated": fixture_updates,
        }
        print(f"[odds] aussportsbetting backfill: {summary}")
        return summary
    finally:
        con.close()
